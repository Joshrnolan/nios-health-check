#!/usr/bin/env python3
"""
Infoblox NIOS Grid Health Audit (Production v27 - single-file edition)
------------------------------------------------------------------------
SINGLE-FILE EDITION: this script is intentionally self-contained (no other
.py files required) so a customer's security team can review one file
top-to-bottom before approving it to run in their environment.

WHAT THIS SCRIPT DOES, SECTION BY SECTION (search for the "SECTION n"
banners below to jump to each part)
------------------------------------------------------------------------
  SECTION 1 - CORE HEALTH CHECK (unchanged since v22-24)
      Read-only WAPI collection that produces the standard 43-column
      Health Check report (.xlsx / .csv / .summary.json). This is the
      part that gets uploaded to the health check portal; its output
      format, column headers (HEADER_43), and cell formatting are NOT
      altered by either optional section below.

  SECTION 2 - OPTIONAL: GRID MEMBER CAPACITY REPORT (--capacity-report)
      Adapted from nios_grid_capacity.py (Pat Vogelsang, MIT License; see
      NOTICE.md). Reuses the SAME authenticated, read-only WAPI session
      opened in Section 1 - no second login. Performs additional GET
      calls against the `capacityreport` WAPI object only. Writes a
      SEPARATE Excel workbook (<run>_grid_capacity.xlsx); never touches
      the Section 1 output files.

  SECTION 3 - OPTIONAL: DNS TOPOLOGY VISUALIZATION (--topology-viz)
      Adapted from ddi_collect.py / ddi_dashboard.html, part of the NIOS
      DDI Dashboard tool written by Bobby Cooper (Infoblox Inc., MIT
      License; see NOTICE.md). Reuses the SAME authenticated, read-only
      WAPI session opened in Section 1. Performs additional GET calls
      against `zone_auth`, `zone_delegated`, `zone_forward`, `zone_stub`,
      `nsgroup`, and `zone_rp` (Response Policy Zones) only. Renders the
      result into a SEPARATE, self-contained
      static HTML file (<run>_topology.html). The relationship-graph
      JavaScript/CSS shipped inside that HTML file is embedded below as
      the TOPOLOGY_HTML_TEMPLATE string constant, clearly marked - it
      only runs later, in the customer's own browser, when they open the
      generated .html file. It never executes on the machine running this
      Python script and never calls out to any network endpoint from
      Python. Never touches the Section 1 output files.

SECURITY / NETWORK SUMMARY (applies to all three sections)
------------------------------------------------------------------------
  * Every WAPI call made by this script is an HTTP GET. The ONLY
    exception is a single HTTP POST to the WAPI `logout` endpoint at the
    very end of a run, which simply releases the session token
    (InfobloxClient.logout) - no object is ever created, modified, or
    deleted on the Grid.
  * The only network endpoint this script talks to from Python is the
    Grid Manager / Grid Master Candidate host or IP supplied interactively
    or via --grid-ip. No telemetry, analytics, or third-party endpoints
    are contacted by this script.
  * Credentials are held in memory only for the duration of the run -
    never written to disk, never logged (see JsonLineFormatter) - and are
    used solely to build the WAPI Basic-Auth session.
  * v25 adds no new outbound network calls beyond additional read-only
    GETs against the same Grid Manager already being audited by Section 1.
  * TOPOLOGY_HTML_TEMPLATE (Section 3) references one third-party CDN
    script (ECharts, from cdn.jsdelivr.net) inside the *generated* HTML
    file, which is loaded by the *customer's browser* when they open that
    file - it is not fetched or executed by this Python script. Review the
    constant directly (search "TOPOLOGY_HTML_TEMPLATE = r\"\"\"") if your
    policy requires vetting that beforehand.

v27 changes vs v26:
  - FIX: Member IP (column F) was blank for grid members that do not run the
    DNS service (e.g. Reporting appliances / IB-V5005, Network Insight /
    Discovery appliances / ND-V906). Root cause: the IP lookup relied solely
    on the member:dns WAPI object, which only returns an entry for members
    with the DNS service present; the intended fallback (parsing an IP out
    of capacityreport._ref) never worked because that _ref ends in the
    member's host name, not an IP address.
  - Added get_member_vip_map(), sourced from member.vip_setting, which is
    universal across every grid member type regardless of which services
    (DNS/DHCP) it runs. Column F now resolves via a 3-tier fallback:
    member:dns -> member.vip_setting -> capacityreport._ref (legacy).
  - Any member for which all three sources come back empty is now logged
    as a WARNING (previously silent) so gaps are visible in the run log.

v26 changes vs v25:
  - Added Name Server Groups, external name servers, and Response Policy
    Zones (zone_rp) to the Section 3 DNS topology visualization.

v25 changes vs v24:
  - Combined what previously shipped as three files (nios_health_check_final.py,
    nios_grid_capacity_module.py, nios_topology_viz.py) into this single
    file, so it can be reviewed and approved as one script.
  - NEW, additive only, both default OFF:
      --capacity-report  Section 2: separate Grid Member Capacity workbook.
      --topology-viz     Section 3: separate DNS topology HTML page.
  - HEADER_43, write_excel(), write_csv(), and the summary.json schema are
    unchanged, so the existing health check portal upload workflow keeps
    working exactly as before.

v24 changes vs v23:
  - FIX: Interactive "Include Member IP Addresses" prompt now actually fires.
         (v23 bug: argparse store_true defaulted the attr to False, not None,
         so the prompt was silently skipped.)
  - Member IP (column F) is now sourced from member:dns?_return_fields=host_name,ipv4addr
    when the user opts in, instead of parsing capacityreport._ref.
  - When user chooses NO, column F is left blank for every row.

v23 changes vs v22:
  - Column C ("grid_uuid") populated from grid.uuid (WAPI v2.14+ / NIOS 9.1.0+)
    with automatic fallback to grid:license_pool_container.lpc_uid on older NIOS.
"""
from __future__ import annotations

import argparse
import csv
import getpass
import hashlib
import json
import logging
import os
from logging.handlers import RotatingFileHandler
import re
from collections import Counter, defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import requests
import urllib3
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# =============================================================================
# SECTION 1 - CORE HEALTH CHECK (unchanged since v22-24)
# Everything from here down to the "END SECTION 1" banner produces the
# standard 43-column Health Check report and is NOT modified by the
# optional Section 2 / Section 3 enhancements below.
# =============================================================================

# ------------------------- Constants -------------------------
APP_NAME            = "nios_health_audit"
DEFAULT_API_VERSION = "v2.12"
REQUEST_TIMEOUT     = 30
RETRY_TOTAL         = 3
RETRY_BACKOFF       = 0.3

ROLE_MAP: Dict[str, str] = {
    "grid master":           "GM",
    "grid master candidate": "GMC",
}

HEADER_43: List[str] = [
    "Customer Name", "Employee Count", "grid_uuid", "Member Serial Number", "Member Role",
    "Member IP", "Member Host Name", "Member Model", "Member Platform", "Member ha Status",
    "Member Operational State", "Member Version", "Member Version History", "Member Protocol",
    "Member Object Count", "Member Enabled Features", "Member License ", "Log Lease Events",
    "Grid Name", "Geo Country Name", "Collected At", "grid wide license con", "Grid Enabled Feature",
    "DNS DNS Scavenging", "DNS Query Response Logging", "DNS Nameserver Groups", "DNS Anycast",
    "DNS DNS Scavenging", "DNS DNS Over HTTPs", "DNS DTC", "DHCP Finger Printing",
    "NIOS File Distribution", "User Name", "DHCP IPv4 Hosts", "DNS DNS Views", "DTC Members Pct",
    "Grid Admin Count", "NIOS Smart Folders", "Member LPS Total", "CPU Usage Ratio",
    "Disk Usage Ratio", "Memeory Usage Ratio", "Member QPS",
]

# ------------------------- Logging -------------------------
class JsonLineFormatter(logging.Formatter):
    def format(self, record: logging.LogRecord) -> str:
        payload = {
            "ts":     datetime.utcnow().isoformat() + "Z",
            "level":  record.levelname,
            "msg":    record.getMessage(),
            "logger": record.name,
        }
        if record.exc_info:
            payload["exception"] = self.formatException(record.exc_info)
        return json.dumps(payload)

def setup_logging(log_path: str, debug: bool = False) -> logging.Logger:
    logger = logging.getLogger("bloxconnect")
    logger.setLevel(logging.DEBUG if debug else logging.INFO)
    logger.handlers.clear()

    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG if debug else logging.INFO)
    ch.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
    logger.addHandler(ch)

    fh = RotatingFileHandler(log_path, maxBytes=10 * 1024 * 1024, backupCount=3)
    fh.setLevel(logging.DEBUG if debug else logging.INFO)
    fh.setFormatter(JsonLineFormatter())
    logger.addHandler(fh)
    return logger

# ------------------------- HTTP Session -------------------------
def make_session(verify_ssl: Any, proxies: Optional[Dict[str, str]] = None) -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=RETRY_TOTAL, read=RETRY_TOTAL, connect=RETRY_TOTAL,
        backoff_factor=RETRY_BACKOFF,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
        raise_on_status=False,
    )
    session.mount("https://", HTTPAdapter(max_retries=retry))
    session.verify = verify_ssl
    if proxies:
        session.proxies.update(proxies)
    return session

# ------------------------- Connection Prompt -------------------------
def gather_connection_info(args: argparse.Namespace) -> Tuple[str, str, str, bool, bool, bool, bool]:
    grid_ip = getattr(args, "grid_ip", "") or ""
    while not grid_ip:
        grid_ip = input("Grid Manager IP/Hostname: ").strip()

    username = getattr(args, "username", "") or ""
    while not username:
        username = input("WAPI Username: ").strip()

    password = getattr(args, "password", "") or ""
    while not password:
        password = getpass.getpass("WAPI Password: ")

    insecure = getattr(args, "insecure", False)
    if not insecure:
        ans = input("Bypass TLS Verification (y/n) [n]: ").strip().lower()
        if ans in ("y", "yes", "1", "true"):
            insecure = True

    # --- IP Address inclusion toggle (v24: fixed so prompt actually fires) ---
    # argparse now uses default=None, so we can distinguish "user passed --include-ip"
    # vs "user didn't pass it, should prompt".
    include_ip_arg = getattr(args, "include_ip", None)
    if include_ip_arg is None:
        ans = input("Include Member IP Addresses in output (y/n) [n]: ").strip().lower()
        include_ip = ans in ("y", "yes", "1", "true")
    else:
        include_ip = bool(include_ip_arg)

    # --- v25.1: Optional enhancement toggles (Section 2 / Section 3) ---
    # Same "default=None means prompt" pattern as include_ip above, so users
    # get an interactive y/n prompt UNLESS they already answered on the CLI
    # via --capacity-report/--no-capacity-report or --topology-viz/--no-topology-viz.
    # This lets the same script be driven either interactively or by flags
    # (e.g. for unattended/scheduled runs), without changing Section 1 output.
    capacity_report_arg = getattr(args, "capacity_report", None)
    if capacity_report_arg is None:
        ans = input("Include Grid Member Database Capacity report (y/n) [n]: ").strip().lower()
        capacity_report = ans in ("y", "yes", "1", "true")
    else:
        capacity_report = bool(capacity_report_arg)

    topology_viz_arg = getattr(args, "topology_viz", None)
    if topology_viz_arg is None:
        ans = input("Include DNS Topology Visualization (y/n) [n]: ").strip().lower()
        topology_viz = ans in ("y", "yes", "1", "true")
    else:
        topology_viz = bool(topology_viz_arg)

    return grid_ip, username, password, insecure, include_ip, capacity_report, topology_viz

def get_latest_wapi_version(
    grid_ip: str, username: str, password: str,
    verify: Any, proxies: Optional[Dict[str, str]], logger: logging.Logger,
) -> str:
    session = make_session(verify, proxies)
    url = f"https://{grid_ip}/wapi/v1.0/?_schema"
    try:
        resp = session.get(url, auth=(username, password), timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        versions = resp.json().get("supported_versions", [])
        if versions:
            def pv(v: str) -> List[int]:
                return [int(x) for x in str(v).lower().lstrip("v").split(".") if x.isdigit()]
            latest = sorted(versions, key=pv)[-1]
            return latest if latest.startswith("v") else f"v{latest}"
    except Exception as e:
        logger.warning(f"Could not auto-detect WAPI version (defaulting to {DEFAULT_API_VERSION}): {e}")
    return DEFAULT_API_VERSION

# ------------------------- Infoblox WAPI Client -------------------------
# SECURITY NOTE: every method on this class issues a WAPI GET (read-only)
# except logout(), which issues one WAPI POST solely to release the session
# token. No object on the Grid is ever created, modified, or deleted.
class InfobloxClient:
    def __init__(
        self, grid_ip: str, username: str, password: str,
        api_version: str = DEFAULT_API_VERSION, verify_ssl: Any = True,
        logger: Optional[logging.Logger] = None, timeout: int = REQUEST_TIMEOUT,
        proxies: Optional[Dict[str, str]] = None,
    ):
        self.base_url = f"https://{grid_ip}/wapi/{api_version}/"
        self.auth     = (username, password)
        self.session  = make_session(verify_ssl, proxies)
        self.timeout  = timeout
        self.logger   = logger or logging.getLogger("bloxconnect")

    def _get(self, endpoint: str, params: Optional[Dict[str, Any]] = None) -> Optional[Any]:
        url = f"{self.base_url}{endpoint}"
        try:
            resp = self.session.get(url, auth=self.auth, params=params, timeout=self.timeout)
            if not (200 <= resp.status_code < 300):
                self.logger.warning(f"GET {endpoint} returned {resp.status_code}")
            resp.raise_for_status()
            data = resp.json()
            return data["result"] if isinstance(data, dict) and "result" in data else data
        except requests.exceptions.SSLError as e:
            self.logger.error(f"TLS error on {endpoint}: {e}")
        except Exception as e:
            self.logger.error(f"GET {endpoint} failed: {e}")
        return None

    def test_connectivity(self) -> bool:
        ok = bool(self._get("grid", {"_return_fields": "name"}))
        if not ok:
            self.logger.error("Connectivity test failed. Check credentials and TLS (--insecure).")
        return ok

    def get_grid_identity(self) -> Dict[str, Any]:
        return (self._get("grid", {"_return_fields": "name"}) or [{}])[0]

    # --------------------------------------------------------------
    # grid_uuid with automatic WAPI-version-aware fallback (v23)
    # --------------------------------------------------------------
    def get_grid_uuid(self, api_ver: str) -> str:
        """
        Returns grid.uuid when running WAPI >= v2.14 (NIOS 9.1.0+).
        For older versions, falls back to grid:license_pool_container.lpc_uid.
        """
        try:
            if wapi_supports_grid_uuid(api_ver):
                data = self._get("grid", {"_return_fields": "uuid"}) or []
                if data and isinstance(data, list):
                    uuid_val = data[0].get("uuid", "")
                    if uuid_val:
                        self.logger.info(f"grid_uuid (uuid) retrieved via WAPI {api_ver}")
                        return uuid_val
                self.logger.warning("grid.uuid not returned; falling back to lpc_uid lookup")

            data = self._get(
                "grid:license_pool_container",
                {"_return_fields": "lpc_uid"},
            ) or []
            if data and isinstance(data, list):
                lpc_uid = data[0].get("lpc_uid", "")
                if lpc_uid:
                    self.logger.info("grid_uuid (lpc_uid fallback) retrieved")
                    return lpc_uid
        except Exception as e:
            self.logger.error(f"get_grid_uuid failed: {e}")
        return "na"

    # --------------------------------------------------------------
    # NEW in v24: member IP map via member:dns?_return_fields=host_name,ipv4addr
    # --------------------------------------------------------------
    def get_member_ipv4_map(self) -> Dict[str, str]:
        """
        Returns { host_name: ipv4addr } for every Grid member, sourced from the
        member:dns endpoint.  Equivalent to:
          curl -k -u USER:PASS \
            "https://GRID/wapi/<ver>/member:dns?_return_fields=host_name,ipv4addr"
        """
        try:
            data = self._get(
                "member:dns",
                {"_return_fields": "host_name,ipv4addr"},
            ) or []
            ip_map: Dict[str, str] = {}
            for item in data:
                host = item.get("host_name")
                ip   = item.get("ipv4addr") or ""
                if host:
                    ip_map[host] = ip
            self.logger.info(f"get_member_ipv4_map: {len(ip_map)} host(s) with IP")
            return ip_map
        except Exception as e:
            self.logger.error(f"get_member_ipv4_map failed: {e}")
            return {}

    # --------------------------------------------------------------
    # NEW in v27: universal member IP fallback via member.vip_setting
    # --------------------------------------------------------------
    def get_member_vip_map(self) -> Dict[str, str]:
        """
        Returns { host_name: vip_setting.address } for every Grid member.

        WHY THIS EXISTS (bug fix, v27):
        member:dns (used by get_member_ipv4_map above) only returns an entry
        for members that actually run the DNS service. Special-purpose grid
        members -- e.g. Reporting appliances (IB-V5005) and Network Insight /
        Discovery appliances (ND-V906) -- do not run DNS, so they never
        appear in member:dns at all and were silently dropped from column F.

        member.vip_setting is the member's base management IP and exists for
        EVERY grid member regardless of which services (DNS/DHCP) it runs, so
        it is used here as a universal second-tier source before the (mostly
        theoretical) capacityreport._ref fallback in get_member_role_and_ip.
        """
        try:
            data = self._get(
                "member",
                {"_return_fields": "host_name,vip_setting"},
            ) or []
            ip_map: Dict[str, str] = {}
            for item in data:
                host = item.get("host_name")
                vip  = item.get("vip_setting") or {}
                addr = vip.get("address") or ""
                if host and addr:
                    ip_map[host] = addr
            self.logger.info(f"get_member_vip_map: {len(ip_map)} host(s) with IP (universal fallback)")
            return ip_map
        except Exception as e:
            self.logger.error(f"get_member_vip_map failed: {e}")
            return {}

    def get_software_version(self) -> str:
        data = self._get("upgradestatus", {"type": "GRID", "_return_fields": "current_version_summary"})
        if data:
            full = data[0].get("current_version_summary", "N/A")
            return full.split("-")[0] if "-" in full else full
        return "N/A"

    def get_global_licenses(self) -> str:
        data = self._get("license:gridwide", {"_return_fields": "type"}) or []
        return ", ".join(sorted(set(l.get("type", "") for l in data if l.get("type"))))

    def get_grid_members(self) -> List[Dict[str, Any]]:
        # ipv4addr excluded on member — causes 400 on WAPI v2.14; we use member:dns instead.
        return self._get("member", {"_return_fields+": "node_info,service_status,host_name,master_candidate"}) or []

    def get_member_role_and_ip(self, host_name: str) -> Tuple[str, str]:
        """
        Authoritative role detection via capacityreport 'role' field.
        Returns (role_label, member_ip).

        NOTE (v24): member_ip returned here is still derived from capacityreport._ref.
        It is used ONLY when the grid-wide member:dns map does not have an entry
        for this host_name (rare fallback).  When include_ip is False at the
        caller, this value is discarded entirely before being written to the row.
        """
        data = self._get("capacityreport", {"name": host_name, "_return_fields": "name,role"})
        if not data:
            self.logger.info(f"capacityreport '{host_name}': no data (offline?) — defaulting to Member")
            return "Member", ""
        entry    = data[0] if isinstance(data, list) else {}
        raw_role = entry.get("role", "").strip().lower()
        role_label = ROLE_MAP.get(raw_role, "Member")
        member_ip = ""
        m = re.search(r":([^/]+)$", entry.get("_ref", ""))
        if m and re.match(r"^\d{1,3}(\.\d{1,3}){3}$", m.group(1)):
            member_ip = m.group(1)
        self.logger.debug(f"capacityreport '{host_name}': role='{raw_role}' -> '{role_label}', ip='{member_ip}'")
        return role_label, member_ip

    def get_licenses_by_hwid(self) -> Dict[str, List[Dict[str, Any]]]:
        data = self._get(
            "member:license",
            {"_return_fields": "type,kind,limit,expiration_status,expiry_date,hwid"},
        ) or []
        result: Dict[str, List[Dict[str, Any]]] = {}
        for lic in data:
            hwid = lic.get("hwid")
            if hwid:
                result.setdefault(hwid, []).append(lic)
            else:
                self.logger.debug(f"member:license has no hwid: {lic.get('_ref')}")
        self.logger.info(f"get_licenses_by_hwid: licenses found for {len(result)} hwid(s)")
        return result

    def get_member_object_count(self, host: str) -> int:
        data = self._get("capacityreport", {"name": host, "_return_fields": "object_counts"})
        if not data:
            return 0
        try:
            return sum(o.get("count", 0) for o in data[0].get("object_counts", []))
        except Exception:
            return 0

    def get_active_dhcp_leases(self, ref: Optional[str]) -> int:
        if not ref:
            return 0
        try:
            res   = self._get("dhcp:statistics", {"statistics_object": ref, "_return_fields": "static_hosts,dynamic_hosts"})
            stats = res[0] if isinstance(res, list) else (res or {})
            return int(stats.get("static_hosts", 0) or 0) + int(stats.get("dynamic_hosts", 0) or 0)
        except Exception:
            return 0

    def get_global_dns_settings(self) -> Tuple[Dict[str, Any], Dict[str, Any]]:
        scav = (self._get("grid:dns", {"_return_fields": "scavenging_settings"})  or [{}])[0]
        logs = (self._get("grid:dns", {"_return_fields": "logging_categories"})   or [{}])[0]
        return scav, logs

    def get_global_dhcp_settings(self) -> Dict[str, Any]:
        props = (self._get("grid:dhcpproperties", {"_return_fields": "log_lease_events"}) or [{}])[0]
        fp    =  self._get("grid:dhcpproperties", {"_return_fields": "enable_fingerprint"})
        props["enable_fingerprint"] = fp[0].get("enable_fingerprint", False) if fp else False
        return props

    def get_dhcp_service_map(self) -> Dict[str, Dict[str, Any]]:
        data = self._get("member:dhcpproperties", {"_return_fields": "host_name,enable_dhcp"}) or []
        return {i["host_name"]: {"ref": i.get("_ref"), "enabled": i.get("enable_dhcp", False)}
                for i in data if i.get("host_name")}

    def get_dns_service_map(self) -> Dict[str, Dict[str, Any]]:
        data = self._get("member:dns", {"_return_fields": "host_name,enable_dns"}) or []
        return {i["host_name"]: {"enabled": i.get("enable_dns", False)}
                for i in data if i.get("host_name")}

    def get_grid_object_counts(self) -> Dict[str, Any]:
        return {
            "views":   len(self._get("view")                or []),
            "has_nsg": bool(self._get("nsgroup")),
            "folders": len(self._get("smartfolder:global") or []),
            "admins":  len(self._get("adminuser")          or []),
        }

    def logout(self) -> None:
        try:
            resp = self.session.post(f"{self.base_url}logout", auth=self.auth, timeout=self.timeout)
            if 200 <= resp.status_code < 300:
                self.logger.info("WAPI session logged out.")
            else:
                self.logger.warning(f"WAPI logout returned {resp.status_code}")
        except Exception as e:
            self.logger.error(f"WAPI logout failed: {e}")

# ------------------------- Output Writers -------------------------
# DO NOT MODIFY: HEADER_43, write_excel(), and write_csv() define the exact
# 43-column format required by the health check portal upload workflow.
def write_excel(rows: List[Dict[str, Any]], path: str, logger: logging.Logger) -> None:
    if not XLSX_AVAILABLE:
        logger.warning("openpyxl not installed — skipping Excel output.")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Health Audit"
    for ci, h in enumerate(HEADER_43, 1):
        ws.cell(row=1, column=ci, value=h)
    for ri, data in enumerate(rows, 2):
        for ci, h in enumerate(HEADER_43, 1):
            val  = data.get("DNS DNS Scavenging_2") if ci == 28 else data.get(h, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci in (40, 41, 42):
                cell.number_format = "0%"
    wb.save(path)

def write_csv(rows: List[Dict[str, Any]], path: str) -> None:
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(HEADER_43)
        for data in rows:
            w.writerow([
                data.get("DNS DNS Scavenging_2") if i == 28 else data.get(h, "")
                for i, h in enumerate(HEADER_43, 1)
            ])

# ------------------------- Helpers -------------------------
def pct_to_ratio(s: str) -> float:
    try:
        return float(s.replace("%", "").strip()) / 100.0
    except Exception:
        return 0.0

def validate_geo(value: Optional[str]) -> str:
    v = (value or "AMS").strip().upper()
    return v if v in {"EMEA", "AMS", "APJ"} else "AMS"

def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

# ------------------------- Version Helpers -------------------------
def parse_wapi_version(v: str) -> Tuple[int, int]:
    """Convert 'v2.14' / '2.14.1' -> (2, 14).  Returns (0, 0) on failure."""
    try:
        parts = [int(x) for x in str(v).lower().lstrip("v").split(".") if x.isdigit()]
        major = parts[0] if len(parts) >= 1 else 0
        minor = parts[1] if len(parts) >= 2 else 0
        return (major, minor)
    except Exception:
        return (0, 0)

def wapi_supports_grid_uuid(api_ver: str) -> bool:
    """grid.uuid is exposed starting at WAPI v2.14 (NIOS 9.1.0)."""
    return parse_wapi_version(api_ver) >= (2, 14)

# =============================================================================
# END SECTION 1 - CORE HEALTH CHECK
# =============================================================================

# =============================================================================
# SECTION 2 - OPTIONAL: GRID MEMBER CAPACITY REPORT  (--capacity-report)
# ------------------------------------------------------------------------
# Adapted from nios_grid_capacity.py (Pat Vogelsang, MIT License; full
# license text reproduced in NOTICE.md).
#
# Network footprint: this section makes ONE additional read-only WAPI GET
# per Grid member, against the `capacityreport` object only, using the
# SAME already-authenticated `client` / WAPI session opened in Section 1.
# It opens no new connections and requires no new credentials.
#
# Output: writes its OWN Excel workbook, <run>_grid_capacity.xlsx, with two
# sheets ("Capacity Summary", "Object Counts"). Never reads or writes the
# Section 1 output files, HEADER_43, write_excel(), or write_csv().
# Runs only when --capacity-report is passed; a failure here is caught and
# logged without aborting or altering the Section 1 report.
# =============================================================================

# Capacity object-type -> bucket, used to estimate DDI vs Active-IP object
# load per member. Mirrors nios_grid_capacity.py's estimate_uddi_objects().
CAPACITY_DDI_TYPES = {
    "A Record/Substitute (A Record) Rule/Substitute (IPv4 Address) Rule",
    "Access Control Item",
    "CNAME Record/Substitute Domain Name/Block/Passthru Rule",
    "DHCP Custom Option",
    "DHCP Range",
    "DNS Traffic Control HTTP Monitor",
    "DNS Traffic Control ICMP Monitor",
    "DNS Traffic Control PDP Monitor",
    "DNS Traffic Control SIP Monitor",
    "DNS Traffic Control SNMP Monitor",
    "Host Alias",
    "Network",
    "Network Container",
    "PTR Record/Substitute (PTR Record) Rule",
    "Router",
    "SVCB Record/Substitute (SVCB Record) Rule",
    "TXT Record/Substitute (TXT Record) Rule",
    "View",
    "Zone",
    "Zone SOA",
}
CAPACITY_ACTIVE_IP_TYPES = {"Fixed Address", "Host", "Host Address"}

CAPACITY_SUMMARY_HEADER = [
    "Member Host Name", "Node HWID", "Node HA Status", "Node Platform",
    "Capacity Role", "Hardware Type", "Max Capacity", "Total Objects",
    "Percent Used", "Estimated DDI Objects", "Estimated Active IP Objects",
    "Estimated Total UDDI Objects", "Report Found",
]


def _estimate_uddi_objects(object_counts: List[Dict[str, Any]]) -> Tuple[int, int, int]:
    """Bucket WAPI object-count entries into DDI / Active-IP totals."""
    ddi_total = 0
    active_ip_total = 0
    for entry in object_counts or []:
        if not isinstance(entry, dict):
            continue
        type_name = entry.get("type_name", "")
        try:
            count = int(entry.get("count", 0) or 0)
        except (TypeError, ValueError):
            count = 0
        if type_name in CAPACITY_DDI_TYPES:
            ddi_total += count
        elif type_name in CAPACITY_ACTIVE_IP_TYPES:
            active_ip_total += count
    return ddi_total, active_ip_total, ddi_total + active_ip_total


def _fetch_capacity_report(client: "InfobloxClient", host_name: str,
                            logger: logging.Logger) -> Optional[Dict[str, Any]]:
    """Read-only GET of the capacityreport WAPI object for one member."""
    try:
        data = client._get(
            "capacityreport",
            {
                "name": host_name,
                "_return_fields": "object_counts,total_objects,hardware_type,max_capacity,name,role,percent_used",
            },
        )
        if isinstance(data, list) and data:
            return data[0]
        if isinstance(data, dict):
            return data
    except Exception as exc:
        logger.warning(f"[grid-capacity] capacityreport lookup failed for {host_name}: {exc}")
    return None


def generate_capacity_report(client: "InfobloxClient", members: List[Dict[str, Any]],
                              out_dir: str, base_name: str, logger: logging.Logger) -> Optional[str]:
    """Collect per-member capacity data and write a standalone Excel workbook.

    Args:
        client: the InfobloxClient already authenticated in Section 1.
        members: the member list already retrieved by Section 1
            (InfobloxClient.get_grid_members()); reused to avoid a second
            WAPI enumeration call.
        out_dir: the run's timestamped output directory (Section 1).
        base_name: the run's base file name; the capacity workbook is
            written alongside the Section 1 report with a
            `_grid_capacity` suffix so it never collides with it.
        logger: the shared logger opened in Section 1.

    Returns:
        The path to the written .xlsx file, or None if openpyxl is
        unavailable.
    """
    if not XLSX_AVAILABLE:
        logger.warning("[grid-capacity] openpyxl not installed \u2014 skipping grid capacity report.")
        return None

    summary_rows: List[List[Any]] = []
    object_rows: List[Dict[str, Any]] = []
    object_type_columns: List[str] = []
    seen_types = set()

    for member in members:
        host_name = member.get("host_name", "N/A")
        node_info = member.get("node_info", []) or [{}]
        capacity = _fetch_capacity_report(client, host_name, logger)
        found = capacity is not None
        role = capacity.get("role", "") if capacity else ""
        hw_type = capacity.get("hardware_type", "") if capacity else ""
        max_cap = capacity.get("max_capacity", "") if capacity else ""
        total_obj = capacity.get("total_objects", "") if capacity else ""
        pct_used = capacity.get("percent_used", "") if capacity else ""
        object_counts = capacity.get("object_counts", []) if capacity else []
        ddi_total, active_ip_total, combined_total = _estimate_uddi_objects(object_counts)

        for node in node_info:
            summary_rows.append([
                host_name,
                node.get("hwid", ""),
                (node.get("ha_status") or "").replace("_", " ").title(),
                node.get("host_platform", node.get("hwtype", "")),
                role, hw_type, max_cap, total_obj, pct_used,
                ddi_total, active_ip_total, combined_total, found,
            ])

        row: Dict[str, Any] = {"host_name": host_name}
        for entry in object_counts or []:
            if not isinstance(entry, dict):
                continue
            type_name = entry.get("type_name", "")
            if not type_name:
                continue
            row[type_name] = entry.get("count", "")
            if type_name not in seen_types:
                seen_types.add(type_name)
                object_type_columns.append(type_name)
        object_rows.append(row)

        logger.info(f"[grid-capacity] {host_name}: role={role or 'n/a'} "
                    f"percent_used={pct_used or 'n/a'} report_found={found}")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Capacity Summary"
    for ci, h in enumerate(CAPACITY_SUMMARY_HEADER, 1):
        ws1.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(summary_rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            if ci == 9 and isinstance(val, (int, float)):  # Percent Used
                cell.number_format = "0.0"
    for ci in range(1, len(CAPACITY_SUMMARY_HEADER) + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = 22

    ws2 = wb.create_sheet("Object Counts")
    obj_header = ["Member Host Name"] + sorted(object_type_columns)
    for ci, h in enumerate(obj_header, 1):
        ws2.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(object_rows, 2):
        ws2.cell(row=ri, column=1, value=row.get("host_name", ""))
        for ci, h in enumerate(obj_header[1:], 2):
            ws2.cell(row=ri, column=ci, value=row.get(h, ""))
    ws2.column_dimensions["A"].width = 24

    out_path = os.path.join(out_dir, f"{base_name}_grid_capacity.xlsx")
    wb.save(out_path)
    logger.info(f"[grid-capacity] wrote {len(summary_rows)} node row(s) to {out_path}")
    return out_path

# =============================================================================
# END SECTION 2 - OPTIONAL: GRID MEMBER CAPACITY REPORT
# =============================================================================

# =============================================================================
# SECTION 3 - OPTIONAL: DNS TOPOLOGY VISUALIZATION  (--topology-viz)
# ------------------------------------------------------------------------
# Adapted from ddi_collect.py / ddi_dashboard.html, part of the NIOS DDI
# Dashboard tool written by Bobby Cooper (Infoblox Inc., MIT License; full
# license text reproduced in NOTICE.md).
#
# Network footprint: this section makes SIX additional read-only WAPI GETs
# total (one each against `zone_auth`, `zone_delegated`, `zone_forward`,
# `zone_stub`, `nsgroup`, `zone_rp`), using the SAME already-authenticated
# `client` / WAPI session opened in Section 1. It opens no new connections
# and requires no new credentials.
#
# Output: writes its OWN self-contained static HTML file,
# <run>_topology.html, with the zone-relationship data embedded directly in
# it (see TOPOLOGY_HTML_TEMPLATE below). That file has no server component:
# a user just opens it in a browser. The only remote resource it references
# is the ECharts charting library from a public CDN (cdn.jsdelivr.net),
# loaded by the BROWSER when the generated file is opened - this Python
# script never fetches or executes that script itself.
# Never reads or writes the Section 1 output files, HEADER_43,
# write_excel(), or write_csv(). Runs only when --topology-viz is passed; a
# failure here is caught and logged without aborting or altering the
# Section 1 report.
# =============================================================================

def _ns_name(entry: Dict[str, Any]) -> str:
    """Extract a display name from a grid/external name-server struct."""
    return entry.get("name") or entry.get("address") or "unknown"


def build_topology(auth: List[Dict[str, Any]], delegated: List[Dict[str, Any]],
                    forward: List[Dict[str, Any]], stub: List[Dict[str, Any]],
                    nsgroups: Optional[List[Dict[str, Any]]] = None,
                    rpz: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    """Build a relationship graph of zones and name servers.

    Captures these relationship layers:
        * hierarchy    - subzone -> parent zone (longest matching suffix)
        * primary      - name server (or name server group) -> zone it is
                         authoritative primary for
        * secondary    - name server -> zone it serves as secondary
        * forwards     - forward zone -> forward target / Grid forwarding member
        * delegated    - delegated zone -> child name server
        * stub         - stub zone -> the server it stubs from
        * nsg-primary  - name server -> name server group it belongs to as primary
        * nsg-secondary- name server -> name server group it belongs to as secondary

    When a zone is served via a Name Server Group (`ns_group`) rather than
    explicit `grid_primary`/`grid_secondaries` entries, NIOS does not
    populate the per-zone name-server fields - the group itself is the
    only place the member name servers are recorded. To keep zones and
    their real name servers linked in that case, this function also
    ingests `nsgroups` (raw `nsgroup` WAPI records) so that the group's
    member servers show up connected to both the group and, transitively,
    every zone that references it.

    `rpz` (raw `zone_rp` WAPI records - Response Policy Zones) are modeled
    the same way as authoritative zones: same hierarchy, name-server-group
    and external-server handling, distinguished only by their zone `kind`
    (RPZ / RPZ FEED / RPZ FIREEYE) for display.

    Returns:
        {"nodes": [...], "links": [...]} ready for a force-directed graph.
    """
    nsgroups = nsgroups or []
    rpz = rpz or []
    nsgroup_by_name: Dict[str, Dict[str, Any]] = {g.get("name"): g for g in nsgroups if g.get("name")}
    nodes: Dict[str, Dict[str, Any]] = {}
    links: List[Dict[str, str]] = []
    degree: Counter = Counter()

    def add_node(node_id: str, name: str, category: str, kind: str = "") -> str:
        if node_id not in nodes:
            nodes[node_id] = {"id": node_id, "name": name, "category": category, "kind": kind or category}
        return node_id

    def add_link(src: str, dst: str, rel: str) -> None:
        links.append({"source": src, "target": dst, "rel": rel})
        degree[src] += 1
        degree[dst] += 1

    all_zones = []
    for z in auth:
        kind = "Authoritative-Secondary" if z.get("external_primaries") else "Authoritative-Primary"
        all_zones.append((z.get("fqdn", ""), z.get("view", "default"), kind))
    for z in delegated:
        all_zones.append((z.get("fqdn", ""), z.get("view", "default"), "Delegated"))
    for z in forward:
        all_zones.append((z.get("fqdn", ""), z.get("view", "default"), "Forward"))
    for z in stub:
        all_zones.append((z.get("fqdn", ""), z.get("view", "default"), "Stub"))
    for z in rpz:
        rpz_type = (z.get("rpz_type") or "").strip()
        kind = f"RPZ ({rpz_type})" if rpz_type else "RPZ"
        all_zones.append((z.get("fqdn", ""), z.get("view", "default"), kind))

    fqdn_to_id: Dict[tuple, str] = {}
    for fqdn, view, kind in all_zones:
        if not fqdn:
            continue
        node_id = f"zone:{view}:{fqdn}"
        add_node(node_id, fqdn, "Zone", kind)
        fqdn_to_id[(view, fqdn)] = node_id

    by_view: defaultdict = defaultdict(list)
    for (view, fqdn) in fqdn_to_id:
        by_view[view].append(fqdn)
    for view, fqdns in by_view.items():
        fqdn_set = set(fqdns)
        for fqdn in fqdns:
            parts = fqdn.split(".")
            for i in range(1, len(parts)):
                parent = ".".join(parts[i:])
                if parent in fqdn_set:
                    add_link(fqdn_to_id[(view, fqdn)], fqdn_to_id[(view, parent)], "hierarchy")
                    break

    # Name Server Groups: add one node per group plus its member servers,
    # independent of whether any zone currently references the group, so
    # the full NSG topology is always visible.
    for gname, g in nsgroup_by_name.items():
        gid = add_node(f"nsg:{gname}", gname, "Name Server Group")
        for m in g.get("grid_primary", []) or []:
            name = _ns_name(m)
            add_link(add_node(f"ns:{name}", name, "Name Server"), gid, "nsg-primary")
        for m in g.get("grid_secondaries", []) or []:
            name = _ns_name(m)
            add_link(add_node(f"ns:{name}", name, "Name Server"), gid, "nsg-secondary")
        for m in g.get("external_primaries", []) or []:
            name = _ns_name(m)
            add_link(add_node(f"ext:{name}", name, "External Primary"), gid, "nsg-primary")
        for m in g.get("external_secondaries", []) or []:
            name = _ns_name(m)
            add_link(add_node(f"extsec:{name}", name, "External Secondary"), gid, "nsg-secondary")

    def link_zone_ns_fields(zid: str, z: Dict[str, Any]) -> None:
        """Link a zone_auth-shaped record (auth zone or RPZ) to its name servers."""
        ns_group_name = z.get("ns_group")
        if ns_group_name and ns_group_name in nsgroup_by_name:
            # Zone is served by a Name Server Group: link the group to the
            # zone; the group's members were already linked to it above.
            gid = add_node(f"nsg:{ns_group_name}", ns_group_name, "Name Server Group")
            add_link(gid, zid, "primary")
        else:
            for m in z.get("grid_primary", []) or []:
                name = _ns_name(m)
                add_link(add_node(f"ns:{name}", name, "Name Server"), zid, "primary")
            for m in z.get("grid_secondaries", []) or []:
                name = _ns_name(m)
                add_link(add_node(f"ns:{name}", name, "Name Server"), zid, "secondary")
        for m in z.get("external_primaries", []) or []:
            name = _ns_name(m)
            add_link(add_node(f"ext:{name}", name, "External Primary"), zid, "primary")
        for m in z.get("external_secondaries", []) or []:
            name = _ns_name(m)
            add_link(add_node(f"extsec:{name}", name, "External Secondary"), zid, "secondary")
        cinfo = z.get("cloud_info") or {}
        if cinfo.get("delegated_member"):
            name = _ns_name(cinfo["delegated_member"])
            add_link(add_node(f"ns:{name}", name, "Name Server"), zid, "primary")

    for z in auth:
        zid = fqdn_to_id.get((z.get("view", "default"), z.get("fqdn", "")))
        if not zid:
            continue
        link_zone_ns_fields(zid, z)
    for z in rpz:
        zid = fqdn_to_id.get((z.get("view", "default"), z.get("fqdn", "")))
        if not zid:
            continue
        link_zone_ns_fields(zid, z)
    for z in forward:
        zid = fqdn_to_id.get((z.get("view", "default"), z.get("fqdn", "")))
        if not zid:
            continue
        for m in z.get("forward_to", []) or []:
            name = _ns_name(m)
            add_link(zid, add_node(f"fwd:{name}", name, "Forward Target"), "forwards")
        for m in z.get("forwarding_servers", []) or []:
            name = _ns_name(m)
            add_link(zid, add_node(f"ns:{name}", name, "Name Server"), "forwards")
    for z in delegated:
        zid = fqdn_to_id.get((z.get("view", "default"), z.get("fqdn", "")))
        if not zid:
            continue
        for m in z.get("delegate_to", []) or []:
            name = _ns_name(m)
            add_link(zid, add_node(f"ns:{name}", name, "Name Server"), "delegated")
    for z in stub:
        zid = fqdn_to_id.get((z.get("view", "default"), z.get("fqdn", "")))
        if not zid:
            continue
        for m in z.get("stub_from", []) or []:
            name = _ns_name(m)
            add_link(zid, add_node(f"ns:{name}", name, "Name Server"), "stub")

    for (view, fqdn), zid in fqdn_to_id.items():
        nodes[zid]["view"] = view
    for node_id, node in nodes.items():
        node["value"] = degree.get(node_id, 0)

    return {"nodes": list(nodes.values()), "links": links}


def _collect_zone_data(client: "InfobloxClient", logger: logging.Logger) -> Dict[str, List[Dict[str, Any]]]:
    """Read-only GETs of the four zone WAPI objects used to build the graph."""
    def safe_get(obj: str, fields: str) -> List[Dict[str, Any]]:
        try:
            data = client._get(obj, {"_return_fields": fields})
            return data or []
        except Exception as exc:
            logger.warning(f"[topology-viz] {obj} lookup failed: {exc}")
            return []

    return {
        "auth": safe_get("zone_auth", "fqdn,view,cloud_info,grid_primary,grid_secondaries,external_primaries,external_secondaries,ns_group"),
        "delegated": safe_get("zone_delegated", "fqdn,view,delegate_to"),
        "forward": safe_get("zone_forward", "fqdn,view,forward_to,forwarding_servers"),
        "stub": safe_get("zone_stub", "fqdn,view,stub_from"),
        "nsgroups": safe_get("nsgroup", "name,grid_primary,grid_secondaries,external_primaries,external_secondaries"),
        "rpz": safe_get("zone_rp", "fqdn,view,rpz_type,grid_primary,grid_secondaries,external_primaries,external_secondaries,ns_group"),
    }


# ---------------------------------------------------------------------------
# TOPOLOGY_HTML_TEMPLATE
# ---------------------------------------------------------------------------
# This constant is the ENTIRE contents of the self-contained HTML file that
# gets written to <run>_topology.html. It is plain HTML/CSS/JavaScript that
# only ever runs later, in a browser, when a user double-clicks that
# generated file - it is NOT executed by this Python script. The one
# external reference inside it is a <script src="https://cdn.jsdelivr.net/...">
# tag for the ECharts charting library, fetched by the browser at that
# later point, not by Python. The literal string "__TOPOLOGY_JSON__" inside
# it is replaced by generate_topology_viz() below with the JSON payload
# collected from this Grid (zone/name-server relationships only - no
# credentials, no IPs beyond what a zone/member name string already reveals).
# ---------------------------------------------------------------------------
TOPOLOGY_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>NIOS DNS Topology</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js"></script>
<style>
:root {
  --bg: #ffffff; --bg-light: #f6f6f6; --bg-mid: #f0f0f0;
  --border: #e5e5e5; --text-pri: #1b1b1b; --text-sec: #727272;
  --primary: #1b1b1b; --primary-hover: #404040; --selected: #e8e8e8;
  --card-shadow: 0 1px 2px rgba(0,0,0,0.06), 0 4px 12px rgba(0,0,0,0.04);
  --n-zone: #6b7280; --n-ns: #2ec27e; --n-ext: #cda515; --n-extsec: #b45309; --n-fwd: #ef601b; --n-view: #9333ea; --n-nsg: #0891b2;
  --e-hierarchy: #8a8f98; --e-primary: #2ec27e; --e-secondary: #343ced;
  --e-forwards: #ef601b; --e-delegated: #dd2ca5; --e-stub: #0891b2;
  --e-nsg-primary: #0891b2; --e-nsg-secondary: #7c9eb2;
  --graph-canvas: #f6f6f6;
}
* { box-sizing: border-box; }
body { font-family: system-ui, -apple-system, "Segoe UI", sans-serif; margin: 0;
  padding: 22px 24px 44px; background: var(--bg); color: var(--text-pri); -webkit-font-smoothing: antialiased; }
h1 { font-weight: 600; font-size: 1.55rem; margin: 0; letter-spacing: -0.01em; }
.shell { max-width: 1200px; margin: 0 auto; }
.subtitle { color: var(--text-sec); font-size: 0.9rem; margin: 4px 0 16px; }
.meta-badges { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 16px; }
.badge { font-size: 0.74rem; padding: 5px 10px; border-radius: 999px; background: var(--bg-light);
  color: var(--text-sec); border: 1px solid var(--border); white-space: nowrap; }
.badge strong { color: var(--text-pri); font-weight: 600; }

.topo-toolbar { display: flex; flex-wrap: wrap; gap: 10px 18px; align-items: center;
  padding: 12px 14px; margin-bottom: 12px; background: var(--bg-light);
  border: 1px solid var(--border); border-radius: 12px; }
.seg { display: inline-flex; border: 1px solid var(--border); border-radius: 8px; overflow: hidden; }
.seg button { border: 0; background: transparent; color: var(--text-sec); padding: 7px 14px;
  font-size: 0.85rem; cursor: pointer; transition: background .15s, color .15s; }
.seg button.active { background: var(--primary); color: var(--bg); }
.seg button:not(.active):hover { background: var(--selected); color: var(--text-pri); }
.spacing { display: inline-flex; align-items: center; gap: 9px; }
.spacing label { font-size: 0.78rem; color: var(--text-sec); white-space: nowrap; }
.spacing input[type=range] { width: 128px; height: 4px; cursor: pointer; accent-color: var(--primary); }
.search { flex: 1; min-width: 150px; }
.search input { width: 100%; padding: 8px 11px; font-size: 0.85rem; border: 1px solid var(--border);
  border-radius: 8px; background: var(--bg); color: var(--text-pri); }
.filters { display: flex; flex-wrap: wrap; gap: 6px; }
.chip { display: inline-flex; align-items: center; gap: 6px; padding: 5px 11px; font-size: 0.78rem;
  border-radius: 999px; cursor: pointer; user-select: none; border: 1px solid var(--border);
  background: var(--bg); color: var(--text-sec); transition: opacity .15s; }
.chip .dash { width: 14px; height: 3px; border-radius: 2px; }
.chip .dot { width: 9px; height: 9px; border-radius: 50%; }
.chip.off { opacity: 0.4; } .chip.active { color: var(--text-pri); }
.legend { display: flex; flex-wrap: wrap; gap: 14px; margin: 0 0 12px; padding-left: 2px; }
.legend span { display: inline-flex; align-items: center; gap: 6px; font-size: 0.78rem; color: var(--text-sec); }
.legend .dot { width: 10px; height: 10px; border-radius: 50%; }
.canvas-wrap { background: var(--graph-canvas); border: 1px solid var(--border);
  border-radius: 12px; box-shadow: var(--card-shadow); overflow: hidden; }
#graph, #tree { width: 100%; height: 640px; } #tree { display: none; }
.footnote { font-size: 0.78rem; color: var(--text-sec); margin-top: 14px; }
.hint-inline { font-size: 0.78rem; color: var(--text-sec); }
.empty { padding: 60px 20px; text-align: center; color: var(--text-sec); }
</style>
</head>
<body>
<div class="shell">
  <h1>NIOS DNS Topology</h1>
  <p class="subtitle">Primary / secondary / forwarder / delegated / stub relationships, collected via WAPI alongside the health check run</p>
  <div class="meta-badges" id="metaBadges"></div>

  <div class="topo-toolbar">
    <div class="seg" id="viewToggle">
      <button data-view="graph" class="active">Relationship Graph</button>
      <button data-view="tree">Zone Hierarchy</button>
    </div>
    <div class="filters" id="viewFilters"></div>
    <div class="filters" id="edgeFilters"></div>
    <div class="spacing" id="spacingCtl">
      <label for="spacing">Node spacing</label>
      <input id="spacing" type="range" min="40" max="240" value="100" aria-label="Adjust spacing between nodes">
    </div>
    <div class="search"><input id="search" type="text" placeholder="Search zone or server&hellip;" autocomplete="off"></div>
  </div>
  <div class="legend" id="legend"></div>
  <div class="canvas-wrap"><div id="graph"></div><div id="tree"></div></div>
  <p class="footnote" id="footnote"></p>
</div>

<script>
var DDI_DATA = __TOPOLOGY_JSON__;

(function () {
  var textPri = '#1b1b1b', textSec = '#727272', border = '#e5e5e5';
  function chart(id, renderer) { var d = document.getElementById(id); return echarts.getInstanceByDom(d) || echarts.init(d, null, { renderer: renderer || 'svg' }); }
  var baseTip = { backgroundColor: '#f6f6f6', borderColor: border, textStyle: { color: textPri, fontSize: 12 } };

  var meta = DDI_DATA.meta || {}, counts = DDI_DATA.zone_counts || {};
  document.getElementById('metaBadges').innerHTML =
    '<span class="badge">Grid: <strong>' + (meta.grid_master || 'n/a') + '</strong></span>' +
    '<span class="badge">Authoritative: <strong>' + (counts.Authoritative || 0) + '</strong></span>' +
    '<span class="badge">Delegated: <strong>' + (counts.Delegated || 0) + '</strong></span>' +
    '<span class="badge">Forward: <strong>' + (counts.Forward || 0) + '</strong></span>' +
    '<span class="badge">Stub: <strong>' + (counts.Stub || 0) + '</strong></span>' +
    '<span class="badge">RPZ: <strong>' + (counts.RPZ || 0) + '</strong></span>' +
    '<span class="badge">Name Server Groups: <strong>' + (counts.NameServerGroups || 0) + '</strong></span>';
  document.getElementById('footnote').textContent =
    'Data source: Infoblox NIOS WAPI (zone_auth, zone_delegated, zone_forward, zone_stub, zone_rp, nsgroup). ' +
    'Collected once when the health check script was run with --topology-viz; reopen this file any time \u2014 no server required.';

  var T = DDI_DATA.topology || { nodes: [], links: [] };
  if (!T.nodes.length) {
    document.querySelector('.canvas-wrap').innerHTML = '<div class="empty">No zone relationships were found on this Grid.</div>';
    return;
  }

  var CAT = [{ name: 'Zone', color: '#6b7280' }, { name: 'Name Server', color: '#2ec27e' }, { name: 'External Primary', color: '#cda515' }, { name: 'External Secondary', color: '#b45309' }, { name: 'Forward Target', color: '#ef601b' }, { name: 'Name Server Group', color: '#0891b2' }];
  var catColor = {}, catIndex = {}; CAT.forEach(function (c, i) { catColor[c.name] = c.color; catIndex[c.name] = i; });
  var REL = [{ key: 'hierarchy', label: 'Subzone of', color: '#8a8f98' }, { key: 'primary', label: 'Primary for', color: '#2ec27e' },
    { key: 'secondary', label: 'Secondary for', color: '#343ced' }, { key: 'forwards', label: 'Forwards to', color: '#ef601b' },
    { key: 'delegated', label: 'Delegated to', color: '#dd2ca5' }, { key: 'stub', label: 'Stub from', color: '#0891b2' },
    { key: 'nsg-primary', label: 'NSG primary member', color: '#0891b2' }, { key: 'nsg-secondary', label: 'NSG secondary member', color: '#7c9eb2' }];
  var relColor = {}, relLabel = {}, active = {}; REL.forEach(function (r) { relColor[r.key] = r.color; relLabel[r.key] = r.label; active[r.key] = true; });
  var nodeById = {}, deg = {}, nodeSpacing = 1.0, viewActive = {};

  function nodeView(n) { return (n && n.view) || 'default'; }
  function syncViews() {
    var seen = {};
    T.nodes.forEach(function (n) { if (n.category === 'Zone') { var vw = nodeView(n); seen[vw] = true; if (!(vw in viewActive)) viewActive[vw] = true; } });
    return Object.keys(viewActive).filter(function (vw) { return seen[vw]; }).sort();
  }

  function prepTopo() {
    deg = {}; T.links.forEach(function (l) { deg[l.source] = (deg[l.source] || 0) + 1; deg[l.target] = (deg[l.target] || 0) + 1; });
    nodeById = {};
    var nodes = T.nodes.map(function (n) {
      var nn = { id: n.id, name: n.name, category: catIndex[n.category] || 0, symbolSize: 8 + Math.min(34, Math.sqrt(deg[n.id] || 0) * 7),
        itemStyle: { color: catColor[n.category] || '#6b7280' }, label: { show: true }, _kind: n.kind || n.category, _cat: n.category, _view: nodeView(n) };
      nodeById[n.id] = nn; return nn;
    });
    return { nodes: nodes };
  }
  function buildVisible(p) {
    var hiddenZone = {};
    T.nodes.forEach(function (n) { if (n.category === 'Zone' && viewActive[nodeView(n)] === false) hiddenZone[n.id] = true; });
    var links = T.links.filter(function (l) { return active[l.rel] && !hiddenZone[l.source] && !hiddenZone[l.target]; });
    var keep = {}; links.forEach(function (l) { keep[l.source] = 1; keep[l.target] = 1; });
    var nodes = p.nodes.filter(function (n) { if (hiddenZone[n.id]) return false; return n._cat === 'Zone' ? true : !!keep[n.id]; });
    var mapped = links.map(function (l) {
      return { source: l.source, target: l.target, _rel: l.rel,
        lineStyle: { color: relColor[l.rel], width: l.rel === 'hierarchy' ? 1.4 : 1, curveness: l.rel === 'hierarchy' ? 0 : 0.12, opacity: 0.7 } };
    });
    return { nodes: nodes, links: mapped };
  }
  function renderGraph() {
    var p = prepTopo(), vis = buildVisible(p);
    chart('graph', 'canvas').setOption({
      tooltip: Object.assign({ formatter: function (x) {
        if (x.dataType === 'edge') { var s = nodeById[x.data.source], t = nodeById[x.data.target]; return (s ? s.name : x.data.source) + ' <b>' + relLabel[x.data._rel] + '</b> ' + (t ? t.name : x.data.target); }
        return '<b>' + x.data.name + '</b><br>' + x.data._kind + '<br><span style="color:' + textSec + '">' + (deg[x.data.id] || 0) + ' connections</span>';
      } }, baseTip),
      legend: [{ data: CAT.map(function (c) { return c.name; }), textStyle: { color: textSec }, top: 6, icon: 'circle', inactiveColor: border }],
      series: [{ type: 'graph', layout: 'force', roam: true, draggable: true,
        categories: CAT.map(function (c) { return { name: c.name, itemStyle: { color: c.color } }; }),
        data: vis.nodes, links: vis.links,
        force: { repulsion: Math.round(160 * nodeSpacing), edgeLength: [Math.round(45 * nodeSpacing), Math.round(130 * nodeSpacing)], gravity: 0.08 / nodeSpacing, friction: 0.16 },
        emphasis: { focus: 'adjacency', label: { show: true }, lineStyle: { width: 2.4, opacity: 1 } },
        label: { position: 'right', color: textPri, fontSize: 11, formatter: function (x) { return x.data.name; } },
        labelLayout: { hideOverlap: true }, scaleLimit: { min: 0.4, max: 6 } }]
    }, true);
  }
  function renderTree() {
    var p = prepTopo();
    var parentOf = {}; T.links.forEach(function (l) { if (l.rel === 'hierarchy') parentOf[l.source] = l.target; });
    var children = {}; Object.keys(parentOf).forEach(function (c) { (children[parentOf[c]] = children[parentOf[c]] || []).push(c); });
    function node(id) { var n = nodeById[id] || { name: id, _kind: '' }; var o = { name: n.name, _kind: n._kind, itemStyle: { color: '#6b7280' }, label: { color: textPri } }; if (children[id]) o.children = children[id].map(node); return o; }
    var byView = {};
    p.nodes.forEach(function (n) {
      if (n._cat !== 'Zone' || parentOf[n.id]) return;
      var vw = n._view || 'default';
      if (viewActive[vw] === false) return;
      (byView[vw] = byView[vw] || []).push(node(n.id));
    });
    var viewBranches = Object.keys(byView).sort().map(function (vw) {
      return { name: 'view: ' + vw, _kind: 'DNS View', itemStyle: { color: '#9333ea' }, label: { color: textPri, fontWeight: 'bold' }, children: byView[vw] };
    });
    chart('tree', 'svg').setOption({
      tooltip: Object.assign({ trigger: 'item', triggerOn: 'mousemove', formatter: function (x) { return '<b>' + x.name + '</b>' + (x.data._kind ? '<br>' + x.data._kind : ''); } }, baseTip),
      series: [{ type: 'tree', data: [{ name: 'DNS Namespace', itemStyle: { color: '#2ec27e' }, label: { color: textPri, fontWeight: 'bold' }, children: viewBranches }],
        top: 24, bottom: 24, left: 90, right: 160, symbolSize: 8, roam: true, initialTreeDepth: 3,
        lineStyle: { color: '#8a8f98', width: 1.2 }, label: { position: 'left', verticalAlign: 'middle', align: 'right', fontSize: 11 },
        leaves: { label: { position: 'right', align: 'left' } }, emphasis: { focus: 'descendant' }, expandAndCollapse: true, animationDuration: 400 }]
    }, true);
  }
  var topoView = 'graph';
  function renderViewChips() {
    var views = syncViews();
    var el = document.getElementById('viewFilters');
    el.innerHTML = views.map(function (vw) {
      var on = viewActive[vw] !== false;
      return '<span class="chip view-chip ' + (on ? 'active' : 'off') + '" data-view-name="' + vw + '"><span class="dot" style="background:#9333ea"></span>view: ' + vw + '</span>';
    }).join('');
    el.querySelectorAll('.view-chip').forEach(function (chip) {
      chip.addEventListener('click', function () {
        var vw = chip.getAttribute('data-view-name');
        viewActive[vw] = !(viewActive[vw] !== false);
        chip.classList.toggle('off', viewActive[vw] === false);
        chip.classList.toggle('active', viewActive[vw] !== false);
        if (topoView === 'graph') renderGraph(); else renderTree();
      });
    });
  }
  function renderTopology() {
    renderViewChips();
    document.getElementById('legend').innerHTML = CAT.map(function (c) { return '<span><span class="dot" style="background:' + c.color + '"></span>' + c.name + '</span>'; }).join('') +
      '<span class="hint-inline">&bull; node size = number of connections &bull; drag to rearrange, scroll to zoom, slider to space nodes &bull; toggle views to show/hide their DNS</span>';
    if (topoView === 'graph') { renderGraph(); chart('graph', 'canvas').resize(); } else { renderTree(); chart('tree', 'svg').resize(); }
  }
  document.getElementById('edgeFilters').innerHTML = REL.map(function (r) { return '<span class="chip active" data-rel="' + r.key + '"><span class="dash" style="background:' + r.color + '"></span>' + r.label + '</span>'; }).join('');
  document.querySelectorAll('.chip[data-rel]').forEach(function (chip) {
    chip.addEventListener('click', function () { var k = chip.getAttribute('data-rel'); active[k] = !active[k]; chip.classList.toggle('off', !active[k]); chip.classList.toggle('active', active[k]); if (topoView === 'graph') renderGraph(); });
  });
  document.querySelectorAll('#viewToggle button').forEach(function (btn) {
    btn.addEventListener('click', function () {
      document.querySelectorAll('#viewToggle button').forEach(function (b) { b.classList.remove('active'); }); btn.classList.add('active');
      topoView = btn.getAttribute('data-view'); var isG = topoView === 'graph';
      document.getElementById('graph').style.display = isG ? 'block' : 'none';
      document.getElementById('tree').style.display = isG ? 'none' : 'block';
      document.getElementById('edgeFilters').style.visibility = isG ? 'visible' : 'hidden';
      document.getElementById('spacingCtl').style.visibility = isG ? 'visible' : 'hidden';
      renderTopology();
    });
  });
  var spacingTimer = null;
  document.getElementById('spacing').addEventListener('input', function (e) {
    nodeSpacing = (+e.target.value) / 100;
    clearTimeout(spacingTimer);
    spacingTimer = setTimeout(function () { if (topoView === 'graph') renderGraph(); }, 130);
  });
  document.getElementById('search').addEventListener('input', function (e) {
    var q = e.target.value.trim().toLowerCase(); var g = chart('graph', 'canvas');
    g.dispatchAction({ type: 'downplay', seriesIndex: 0 }); if (!q) return;
    var vis = buildVisible(prepTopo()); vis.nodes.forEach(function (n, i) { if (n.name.toLowerCase().indexOf(q) !== -1) g.dispatchAction({ type: 'highlight', seriesIndex: 0, dataIndex: i }); });
  });

  renderTopology();
  window.addEventListener('resize', function () { chart('graph', 'canvas').resize(); chart('tree', 'svg').resize(); });
})();
</script>
</body>
</html>
"""


def generate_topology_viz(client: "InfobloxClient", out_dir: str, base_name: str,
                           logger: logging.Logger) -> Optional[str]:
    """Collect zone relationships and write a self-contained topology HTML page.

    Args:
        client: the InfobloxClient already authenticated in Section 1.
        out_dir: the run's timestamped output directory (Section 1).
        base_name: the run's base file name; the topology page is written
            alongside the Section 1 report with a `_topology` suffix.
        logger: the shared logger opened in Section 1.

    Returns:
        The path to the written .html file.
    """
    zones = _collect_zone_data(client, logger)
    topology = build_topology(zones["auth"], zones["delegated"], zones["forward"], zones["stub"],
                               zones["nsgroups"], zones["rpz"])

    grid_name = "NIOS Grid"
    try:
        grid_name = client.base_url.split("//")[-1].split("/")[0]
    except Exception:
        pass

    payload = {
        "meta": {"grid_master": grid_name},
        "topology": topology,
        "zone_counts": {
            "Authoritative": len(zones["auth"]),
            "Delegated": len(zones["delegated"]),
            "Forward": len(zones["forward"]),
            "Stub": len(zones["stub"]),
            "NameServerGroups": len(zones["nsgroups"]),
            "RPZ": len(zones["rpz"]),
        },
    }

    html = TOPOLOGY_HTML_TEMPLATE.replace("__TOPOLOGY_JSON__", json.dumps(payload))
    out_path = os.path.join(out_dir, f"{base_name}_topology.html")
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(html)

    logger.info(f"[topology-viz] wrote {len(topology['nodes'])} node(s), "
                f"{len(topology['links'])} link(s) to {out_path}")
    return out_path

# =============================================================================
# END SECTION 3 - OPTIONAL: DNS TOPOLOGY VISUALIZATION
# =============================================================================

# =============================================================================
# MAIN ORCHESTRATION
# ------------------------------------------------------------------------
# Runs Section 1 (the standard health check) end-to-end, then optionally
# invokes Section 2 and/or Section 3 using the SAME authenticated `client`
# from Section 1 if the corresponding flag was passed. The Section 1
# output block (write_excel/write_csv/summary.json) is emitted first and
# is completely unaffected by whether the optional sections run or fail.
# =============================================================================
def collect_and_report(args: argparse.Namespace) -> None:
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = f"{APP_NAME}_{ts}"
    os.makedirs(out_dir, exist_ok=True)
    log_path = os.path.join(out_dir, args.log or f"{APP_NAME}_{ts}.log.jsonl")
    logger   = setup_logging(log_path, debug=args.debug)

    print("""
============================================================
INFOBLOX HEALTH CHECK DATA COLLECTION
============================================================
Please provide the following information:
------------------------------------------------------------""")

    grid_ip, username, password, insecure, include_ip, capacity_report, topology_viz = gather_connection_info(args)
    verify_ssl = not insecure

    customer  = args.customer or input("Customer Name: ").strip() or "General"
    employees = str(args.employees) if args.employees is not None else (input("Employee Count [100]: ").strip() or "100")
    geo       = validate_geo(args.geo) if args.geo else validate_geo(input("Geo Country Name (EMEA, AMS, or APJ) [AMS]: ").strip())
    user_name = args.user or input("User/SE Name: ").strip() or "SE"

    print("------------------------------------------------------------\n")
    print(f"Include Member IP Addresses in output: {'YES' if include_ip else 'NO'}")
    print(f"Include Grid Member Database Capacity report: {'YES' if capacity_report else 'NO'}")
    print(f"Include DNS Topology Visualization: {'YES' if topology_viz else 'NO'}\n")

    if not verify_ssl or args.silent_warnings:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    provided_ver = getattr(args, "api_version", "") or ""
    if provided_ver:
        api_ver = provided_ver
        logger.info(f"Using provided API version: {api_ver}")
    else:
        logger.info(f"Connecting to {grid_ip} to auto-detect latest WAPI version.")
        api_ver = get_latest_wapi_version(grid_ip, username, password, verify_ssl, None, logger)
        logger.info(f"Auto-detected WAPI version: {api_ver}")
    print(f"Connecting to Infoblox Grid at {grid_ip} (API {api_ver}).")

    client = InfobloxClient(grid_ip=grid_ip, username=username, password=password,
                            api_version=api_ver, verify_ssl=verify_ssl, logger=logger)

    if not client.test_connectivity():
        return

    print("\n================================================================================")
    print("DATA COLLECTION PHASE")
    print("================================================================================\n")

    grid_name = client.get_grid_identity().get("name", "N/A")
    print(f"Grid Name: {grid_name}")

    ver = client.get_software_version()
    print(f"Grid Version: {ver}")

    grid_uuid = client.get_grid_uuid(api_ver)
    if wapi_supports_grid_uuid(api_ver):
        print(f"Grid UUID: {grid_uuid}")
    else:
        print(f"Grid UUID (license_UID fallback for WAPI {api_ver}): {grid_uuid}")

    grid_lics = client.get_global_licenses()
    members = client.get_grid_members()
    print(f"Found {len(members)} member(s)")

    licenses_by_hwid = client.get_licenses_by_hwid()
    print(f"Found licenses for {len(licenses_by_hwid)} unique serial number(s)")

    # --- NEW in v24: only call member:dns IP endpoint if the user opted in ---
    # --- FIX in v27: member:dns only covers DNS-serving members, so a second
    #     universal source (member.vip_setting) is also collected to catch
    #     Reporting / Network Insight / other non-DNS appliance members. ---
    member_ip_map: Dict[str, str] = {}
    member_vip_map: Dict[str, str] = {}
    if include_ip:
        member_ip_map = client.get_member_ipv4_map()
        print(f"Retrieved IPv4 addresses for {len(member_ip_map)} member(s) via member:dns")
        member_vip_map = client.get_member_vip_map()
        print(f"Retrieved IPv4 addresses for {len(member_vip_map)} member(s) via member.vip_setting (universal fallback)")

    dns_scav, _  = client.get_global_dns_settings()
    _, dns_log   = client.get_global_dns_settings()
    dhcp_global  = client.get_global_dhcp_settings()
    dhcp_map     = client.get_dhcp_service_map()
    grid_counts  = client.get_grid_object_counts()
    dns_map      = client.get_dns_service_map()

    print(f"Found {grid_counts['views']} DNS view(s)")
    print(f"Nameserver Groups configured: {grid_counts['has_nsg']}")
    print(f"Found {grid_counts['folders']} smart folder(s)")
    print(f"Found {grid_counts['admins']} admin user(s)")

    print("\n================================================================================")
    print("PROCESSING MEMBER DATA")
    print("================================================================================\n")

    results: List[Dict[str, Any]] = []
    for idx, member in enumerate(members, 1):
        h_name = member.get("host_name", "N/A")
        print(f"[{idx}/{len(members)}] Processing member: {h_name}")

        # --- Role (always) + fallback IP from capacityreport._ref ---
        role_label, cap_report_ip = client.get_member_role_and_ip(h_name)

        # --- Column-F IP value is controlled by the user's opt-in (v24) ---
        # FIX (v27): 3-tier fallback so non-DNS appliance members (Reporting,
        # Network Insight/Discovery, etc.) no longer come back blank.
        #   1. member:dns ipv4addr           - only present for DNS-serving members
        #   2. member.vip_setting.address    - universal, present for every member
        #   3. capacityreport._ref parse      - legacy last resort (rarely matches)
        if include_ip:
            member_ip_value = (
                member_ip_map.get(h_name)
                or member_vip_map.get(h_name)
                or cap_report_ip
                or ""
            )
            if not member_ip_value:
                logger.warning(f"No IP address found for member '{h_name}' via member:dns, "
                                f"member.vip_setting, or capacityreport._ref.")
        else:
            member_ip_value = ""

        print(f"  - Role: {role_label} | IP: {member_ip_value or '(not included)'}")

        obj_count = client.get_member_object_count(h_name)
        print(f"  - Object count: {obj_count}")

        base_protocols: set = set()
        if dns_map.get(h_name, {}).get("enabled"):
            base_protocols.add("dns")
        if dhcp_map.get(h_name, {}).get("enabled"):
            base_protocols.add("dhcp")

        dhcp_hosts = client.get_active_dhcp_leases(dhcp_map.get(h_name, {}).get("ref"))

        node_info_data = member.get("node_info", [])
        if not isinstance(node_info_data, list) or not node_info_data:
            node_info_data = [{}]

        is_ha = len(node_info_data) == 2
        if is_ha:
            print(f"  - HA Pair detected: processing Active and Passive nodes.")

        for node_idx, node in enumerate(node_info_data):
            node_hwid = node.get("hwid", "")
            if is_ha:
                print(f"    - Node {node_idx+1}: ha_status={node.get('ha_status','N/A').upper()}, hwid={node_hwid or 'N/A'}")

            node_lics   = licenses_by_hwid.get(node_hwid, [])
            lic_types   = [l.get("type", "").lower() for l in node_lics]
            features: set = set(lic_types)
            if any("response policy" in x or "rpz" in x for x in lic_types):
                features.add("rpz")
            if any("threat" in x or "analytics" in x for x in lic_types):
                features.add("threat insight")
            license_str = ", ".join(sorted(set(l.get("type", "") for l in node_lics if l.get("type"))))

            perf: Dict[str, Any] = {"cpu": "0%", "disk": "0%", "mem": "0%", "doh": False}
            node_protocols = set(base_protocols)
            for svc in node.get("service_status", []):
                s_name = svc.get("service")
                desc   = svc.get("description", "")
                status = svc.get("status")
                try:
                    if s_name == "CPU_USAGE":
                        perf["cpu"]  = desc.split(":")[-1].strip()
                    elif s_name == "DISK_USAGE":
                        perf["disk"] = desc.split("%")[0].strip() + "%"
                    elif s_name == "MEMORY":
                        perf["mem"]  = desc.split("%")[0].strip() + "%"
                    elif s_name == "DOT_DOH":
                        perf["doh"]  = (status == "WORKING")
                    if status in ("WORKING", "Running") and s_name in ("NTP", "TFTP", "HTTP", "FTP", "SNMP"):
                        node_protocols.add(s_name.lower())
                except Exception:
                    pass

            results.append({
                "Customer Name":              customer,
                "Employee Count":             employees,
                "grid_uuid":                  grid_uuid,
                "Member Serial Number":       node_hwid or "N/A",
                "Member Role":                role_label,
                "Member IP":                  member_ip_value,   # v24: blank when user opts out
                "Member Host Name":           h_name,
                "Member Model":               node.get("hwtype", "N/A"),
                "Member Platform":            (node.get("host_platform") or node.get("hwplatform") or node.get("hypervisor") or "N/A"),
                "Member ha Status":           node.get("ha_status", "Not Configured").replace("_", " ").title(),
                "Member Operational State":   "Running",
                "Member Version":             ver,
                "Member Version History":     "",
                "Member Protocol":            ", ".join(sorted(node_protocols)),
                "Member Object Count":        obj_count,
                "Member Enabled Features":    ", ".join(sorted(features)),
                "Member License ":            license_str,
                "Log Lease Events":           dhcp_global.get("log_lease_events", False),
                "Grid Name":                  grid_name,
                "Geo Country Name":           geo,
                "Collected At":               datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "grid wide license con":      grid_lics,
                "Grid Enabled Feature":       "",
                "DNS DNS Scavenging":         dns_scav.get("scavenging_settings", {}).get("enable_scavenging", False),
                "DNS Query Response Logging": dns_log.get("logging_categories", {}).get("log_responses", False),
                "DNS Nameserver Groups":      grid_counts["has_nsg"],
                "DNS Anycast":                False,
                "DNS DNS Scavenging_2":       dns_scav.get("scavenging_settings", {}).get("enable_scavenging", False),
                "DNS DNS Over HTTPs":         perf["doh"],
                "DNS DTC":                    False,
                "DHCP Finger Printing":       dhcp_global.get("enable_fingerprint", False),
                "NIOS File Distribution":     False,
                "User Name":                  user_name,
                "DHCP IPv4 Hosts":            dhcp_hosts,
                "DNS DNS Views":              grid_counts["views"],
                "DTC Members Pct":            "Not used",
                "Grid Admin Count":           grid_counts["admins"],
                "NIOS Smart Folders":         grid_counts["folders"],
                "Member LPS Total":           "",
                "CPU Usage Ratio":            pct_to_ratio(perf["cpu"]),
                "Disk Usage Ratio":           pct_to_ratio(perf["disk"]),
                "Memeory Usage Ratio":        pct_to_ratio(perf["mem"]),
                "Member QPS":                 "",
            })

    if not results:
        logger.error("No data collected; skipping output.")
        return

    print("\n================================================================================")
    print("GENERATING OUTPUT FILES")
    print("================================================================================\n")

    base_name     = f"{APP_NAME}_{ts}"
    hashes: Dict[str, str] = {}
    created_files: List[str] = []

    if args.format in ("excel", "both") and XLSX_AVAILABLE:
        ep = os.path.join(out_dir, base_name + ".xlsx")
        write_excel(results, ep, logger)
        hashes[base_name + ".xlsx"] = sha256_file(ep)
        created_files.append(ep)
        print(f"Created: {ep}")

    if args.format in ("csv", "both"):
        cp = os.path.join(out_dir, base_name + ".csv")
        write_csv(results, cp)
        hashes[base_name + ".csv"] = sha256_file(cp)
        created_files.append(cp)
        print(f"Created: {cp}")

    # ------------------------------------------------------------------
    # OPTIONAL ENHANCEMENTS (Section 2 / Section 3) — additive only.
    # Everything above this point (HEADER_43, write_excel(), write_csv(),
    # the .xlsx/.csv files, and the summary.json written below) is
    # completely unchanged, so the standard health check portal upload
    # workflow keeps working exactly as before. These two blocks only run
    # when explicitly requested via --capacity-report / --topology-viz,
    # reuse the already-authenticated `client` session opened above, and
    # write their own separate files. A failure in either one is logged
    # but never aborts or alters the Section 1 report.
    # ------------------------------------------------------------------
    if capacity_report:
        print("\n================================================================================")
        print("OPTIONAL: GRID MEMBER CAPACITY REPORT (Section 2)")
        print("================================================================================\n")
        try:
            cap_path = generate_capacity_report(client, members, out_dir, base_name, logger)
            if cap_path:
                created_files.append(cap_path)
                print(f"Created: {cap_path}")
        except Exception as e:
            logger.error(f"Grid capacity report failed: {e}")

    if topology_viz:
        print("\n================================================================================")
        print("OPTIONAL: DNS TOPOLOGY VISUALIZATION (Section 3)")
        print("================================================================================\n")
        try:
            topo_path = generate_topology_viz(client, out_dir, base_name, logger)
            if topo_path:
                created_files.append(topo_path)
                print(f"Created: {topo_path}")
        except Exception as e:
            logger.error(f"Topology visualization failed: {e}")

    summary = {
        "grid_ip": grid_ip, "api_version": api_ver, "grid_name": grid_name,
        "grid_uuid": grid_uuid, "include_ip": include_ip,
        "capacity_report": capacity_report, "topology_viz": topology_viz,
        "member_count": len(members), "row_count": len(results),
        "views": grid_counts["views"], "admins": grid_counts["admins"],
        "folders": grid_counts["folders"], "has_nsg": grid_counts["has_nsg"],
        "collected_at": datetime.now().isoformat(), "customer": customer,
        "geo": geo, "user": user_name, "format": args.format,
        "log_file": log_path, "hashes": hashes,
    }
    sp = os.path.join(out_dir, base_name + ".summary.json")
    with open(sp, "w") as sf:
        json.dump(summary, sf, indent=2)

    print(f"\nTotal logical members : {len(members)}")
    print(f"Total physical rows   : {len(results)}")
    print(f"Output directory      : {os.path.abspath(out_dir)}")

    print("\nLogging out of WAPI session.")
    client.logout()

# =============================================================================
# CLI
# =============================================================================
def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Infoblox NIOS Health Audit (v25, single-file edition)")
    p.add_argument("--grid-ip")
    p.add_argument("--customer")
    p.add_argument("--employees",       type=int)
    p.add_argument("--geo",             choices=["EMEA", "AMS", "APJ"])
    p.add_argument("--user")
    p.add_argument("--format",          choices=["excel", "csv", "both"], default="both")
    p.add_argument("--log")
    p.add_argument("--insecure",        action="store_true")
    p.add_argument("--silent-warnings", action="store_true")
    p.add_argument("--username")
    p.add_argument("--password")
    p.add_argument("--api-version")
    # v24 fix: default=None so the interactive prompt runs when --include-ip/--no-include-ip
    # is NOT passed on the CLI.  store_true alone defaults to False, which suppressed the prompt.
    p.add_argument("--include-ip",    dest="include_ip", action="store_true",
                   default=None,
                   help="Include Member IP addresses in the output (skips interactive prompt).")
    p.add_argument("--no-include-ip", dest="include_ip", action="store_false",
                   help="Exclude Member IP addresses (skips interactive prompt).")
    p.add_argument("--debug",         action="store_true")
    # --- v25 / v25.1: optional, additive-only enhancements (Section 2 / Section 3) ---
    # default=None (not False) on both, mirroring --include-ip, so
    # gather_connection_info() can tell "user didn't answer yet" (None, prompt
    # interactively) apart from "user explicitly said no" (False, skip the
    # prompt). Passing either flag on the command line always skips its
    # prompt, which is what unattended/scheduled runs need.
    p.add_argument("--capacity-report", dest="capacity_report", action="store_true",
                   default=None,
                   help="Also generate a separate Grid Member Capacity Excel report "
                        "(Section 2, adapted from Pat Vogelsang's nios_grid_capacity.py). "
                        "Skips the interactive prompt. Does not modify the standard "
                        "43-column health check report.")
    p.add_argument("--no-capacity-report", dest="capacity_report", action="store_false",
                   help="Skip the Grid Member Capacity report and its interactive prompt.")
    p.add_argument("--topology-viz", dest="topology_viz", action="store_true",
                   default=None,
                   help="Also generate a self-contained DNS topology visualization "
                        "HTML file (Section 3, adapted from Bobby Cooper's NIOS DDI "
                        "Dashboard) showing primary/secondary/forwarder/delegated/stub "
                        "relationships. Skips the interactive prompt. Does not modify "
                        "the standard 43-column health check report.")
    p.add_argument("--no-topology-viz", dest="topology_viz", action="store_false",
                   help="Skip the DNS Topology Visualization and its interactive prompt.")
    return p

if __name__ == "__main__":
    collect_and_report(build_arg_parser().parse_args())
